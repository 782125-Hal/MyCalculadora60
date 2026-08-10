"""
Importación de datos desde CSV y Excel.

El módulo es puro: lee y valida, pero no toca la base. Las vistas deciden qué
hacer con el resultado, lo que permite mostrar una vista previa antes de
escribir nada — importante cuando lo que entra son cifras de dinero.

Cada tipo de importación declara sus columnas y cómo validar una fila. Los
encabezados se normalizan (minúsculas, sin acentos ni espacios), así que
"Fecha de Pago", "fecha_pago" y "FECHA PAGO" son la misma columna.
"""
import csv
import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

MAX_FILAS = 1000
FORMATOS_FECHA = (
    '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y',
    '%d/%m/%y', '%Y/%m/%d', '%d.%m.%Y',
)


def normalizar(texto):
    """'Fecha de Pago' -> 'fecha_de_pago'. Sin acentos, para tolerar encabezados."""
    if texto is None:
        return ''
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    texto = texto.replace(' ', '_').replace('-', '_')
    return ''.join(c for c in texto if c.isalnum() or c == '_')


# Sinónimos de encabezado. Se aplican tras normalizar, y sólo si la columna
# canónica no vino ya en el archivo.
ALIAS_COLUMNAS = {
    'fecha_de_pago': 'fecha', 'fecha_pago': 'fecha', 'fecha_del_pago': 'fecha',
    'fecha_de_movimiento': 'fecha', 'fecha_movimiento': 'fecha', 'fecha_operacion': 'fecha',
    'fecha_de_compra': 'fecha_compra', 'fecha_de_inicio': 'fecha_inicio',
    'fecha_del_prestamo': 'fecha_inicio', 'inicio': 'fecha_inicio',
    'importe': 'monto', 'cantidad': 'monto', 'monto_pago': 'monto',
    'monto_del_pago': 'monto', 'abono': 'monto', 'monto_invertido': 'monto',
    'concepto_del_pago': 'descripcion', 'detalle': 'descripcion',
    'observaciones': 'descripcion', 'comentario': 'descripcion',
    'cliente': 'contraparte', 'nombre_cliente': 'contraparte', 'acreedor': 'contraparte',
    'deudor': 'contraparte', 'vendedor': 'contraparte',
    'instrumento': 'nombre', 'nombre_del_instrumento': 'nombre',
    'tasa_anual': 'tasa', 'tasa_de_interes': 'tasa', 'tasa_interes_anual': 'tasa',
    'interes': 'tasa',
    'cuota': 'pago', 'pago_mensual': 'pago', 'mensualidad': 'pago',
    'plazo_meses': 'plazo', 'periodos': 'plazo', 'plazo_en_periodos': 'plazo',
    'tipo_pago': 'frecuencia', 'periodicidad': 'frecuencia',
    'valor_actual': 'valor', 'valor_hoy': 'valor',
}


def aplicar_alias(fila):
    """Traduce encabezados sinónimos a la columna canónica, sin pisar la original."""
    for origen, destino in ALIAS_COLUMNAS.items():
        if origen in fila and destino not in fila:
            fila[destino] = fila[origen]
    return fila


def leer_tabla(archivo, nombre):
    """
    Devuelve (filas, error). Cada fila es un dict con claves normalizadas.

    Acepta .csv (con separador , o ;) y .xlsx. El archivo llega como el objeto
    subido por Django, así que se lee en memoria: MAX_FILAS acota el gasto.
    """
    nombre = (nombre or '').lower()
    try:
        if nombre.endswith('.xlsx') or nombre.endswith('.xlsm'):
            return _leer_excel(archivo)
        if nombre.endswith('.csv') or nombre.endswith('.txt'):
            return _leer_csv(archivo)
        return [], 'Formato no reconocido. Usa un archivo .csv o .xlsx.'
    except Exception as exc:  # archivo corrupto, hoja vacía, encoding raro…
        return [], f'No se pudo leer el archivo: {exc}'


def _leer_csv(archivo):
    crudo = archivo.read()
    if isinstance(crudo, bytes):
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                texto = crudo.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return [], 'No se pudo decodificar el archivo. Guárdalo como UTF-8.'
    else:
        texto = crudo

    muestra = texto[:4096]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=',;\t')
    except csv.Error:
        dialecto = csv.excel  # una sola columna, o separador poco común

    lector = csv.DictReader(io.StringIO(texto), dialect=dialecto)
    if not lector.fieldnames:
        return [], 'El archivo no tiene encabezados.'

    filas = []
    for i, cruda in enumerate(lector, start=2):  # 1 es el encabezado
        if i - 1 > MAX_FILAS:
            return [], f'El archivo excede {MAX_FILAS} filas. Divídelo en partes.'
        fila = {normalizar(k): (v.strip() if isinstance(v, str) else v)
                for k, v in cruda.items() if k}
        if any(v not in (None, '') for v in fila.values()):
            fila['_linea'] = i
            filas.append(aplicar_alias(fila))
    return filas, None


def _leer_excel(archivo):
    from openpyxl import load_workbook
    libro = load_workbook(archivo, read_only=True, data_only=True)
    hoja = libro.active
    filas_iter = hoja.iter_rows(values_only=True)

    try:
        encabezados = [normalizar(c) for c in next(filas_iter)]
    except StopIteration:
        return [], 'La hoja está vacía.'

    filas = []
    for i, cruda in enumerate(filas_iter, start=2):
        if i - 1 > MAX_FILAS:
            return [], f'El archivo excede {MAX_FILAS} filas. Divídelo en partes.'
        if all(c in (None, '') for c in cruda):
            continue
        fila = {}
        for encabezado, valor in zip(encabezados, cruda):
            if not encabezado:
                continue
            # openpyxl devuelve datetime para celdas con formato de fecha
            if isinstance(valor, datetime):
                valor = valor.date()
            fila[encabezado] = valor
        fila['_linea'] = i
        filas.append(aplicar_alias(fila))
    libro.close()
    return filas, None


# --------------------------------------------------------------------------
# Conversores. Cada uno lanza ValueError con un mensaje para el usuario final.
# --------------------------------------------------------------------------

def a_fecha(valor, campo='fecha'):
    if isinstance(valor, date):
        return valor
    if isinstance(valor, datetime):
        return valor.date()
    texto = str(valor or '').strip()
    if not texto:
        raise ValueError(f'Falta {campo}.')
    for formato in FORMATOS_FECHA:
        try:
            parsed = datetime.strptime(texto, formato).date()
            break
        except ValueError:
            continue
    else:
        raise ValueError(f'{campo}: "{texto}" no es una fecha reconocible (usa AAAA-MM-DD).')
    # Un año de dos cifras mal capturado produce fechas absurdas que después
    # descuadran los períodos; se rechaza en la puerta.
    if parsed.year < 1900 or parsed.year > 2200:
        raise ValueError(f'{campo}: el año {parsed.year} parece mal capturado.')
    return parsed


def _separar_miles_y_decimales(texto):
    """
    Interpreta '.' y ',' sin saber el idioma del archivo.

    Excel en español escribe 9,75 (coma decimal) y en inglés 9.75, y ambos usan
    el otro símbolo para los miles. La regla: si aparecen los dos, manda el que
    esté más a la derecha; si sólo hay uno y le siguen 1 o 2 dígitos hasta el
    final, es decimal; si le siguen exactamente 3, es separador de miles.
    """
    tiene_punto, tiene_coma = '.' in texto, ',' in texto

    if tiene_punto and tiene_coma:
        decimal = '.' if texto.rfind('.') > texto.rfind(',') else ','
        miles = ',' if decimal == '.' else '.'
        return texto.replace(miles, '').replace(decimal, '.')

    simbolo = '.' if tiene_punto else (',' if tiene_coma else None)
    if simbolo is None:
        return texto

    partes = texto.split(simbolo)
    if len(partes) == 2 and 1 <= len(partes[1]) <= 2:
        return partes[0] + '.' + partes[1]      # 9,75 -> 9.75
    return texto.replace(simbolo, '')            # 1,234 / 1.234 -> 1234


def a_decimal(valor, campo='monto', minimo=None):
    if valor is None or str(valor).strip() == '':
        raise ValueError(f'Falta {campo}.')
    texto = str(valor).strip().replace('$', '').replace(' ', '').replace('\xa0', '')
    texto = _separar_miles_y_decimales(texto)
    try:
        numero = Decimal(texto)
    except InvalidOperation:
        raise ValueError(f'{campo}: "{valor}" no es un número.')
    if minimo is not None and numero < minimo:
        raise ValueError(f'{campo}: debe ser mayor o igual a {minimo}.')
    return numero.quantize(Decimal('0.01'))


def a_entero(valor, campo, minimo=None):
    if valor is None or str(valor).strip() == '':
        raise ValueError(f'Falta {campo}.')
    try:
        numero = int(Decimal(str(valor).strip()))
    except (InvalidOperation, ValueError):
        raise ValueError(f'{campo}: "{valor}" no es un número entero.')
    if minimo is not None and numero < minimo:
        raise ValueError(f'{campo}: debe ser mayor o igual a {minimo}.')
    return numero


def a_opcion(valor, campo, mapa, defecto=None):
    """Traduce sinónimos del usuario al valor interno. 'Pago', 'pago', 'PAGO' -> 'pago'."""
    texto = normalizar(valor)
    if not texto:
        if defecto is not None:
            return defecto
        raise ValueError(f'Falta {campo}.')
    if texto in mapa:
        return mapa[texto]
    opciones = ', '.join(sorted(set(mapa.keys())))
    raise ValueError(f'{campo}: "{valor}" no es válido. Opciones: {opciones}.')


def texto_opcional(valor, maximo=200):
    return str(valor or '').strip()[:maximo]


# --------------------------------------------------------------------------
# Definición de cada importación
# --------------------------------------------------------------------------

TIPOS_MOVIMIENTO_PRESTAMO = {
    'pago': 'pago', 'abono': 'pago', 'pagos': 'pago',
    'incremento': 'incremento_capital', 'incremento_capital': 'incremento_capital',
    'cargo': 'incremento_capital', 'capital': 'incremento_capital',
}

TIPOS_MOVIMIENTO_INVERSION = {
    'aportacion': 'aportacion', 'aporte': 'aportacion', 'deposito': 'aportacion',
    'retiro': 'retiro', 'retiros': 'retiro',
    'rendimiento': 'rendimiento', 'interes': 'rendimiento', 'intereses': 'rendimiento',
}

ROLES = {
    'prestamo': 'prestamo', 'prestamos': 'prestamo', 'otorgado': 'prestamo',
    'me_deben': 'prestamo', 'presto': 'prestamo',
    'deuda': 'deuda', 'deudas': 'deuda', 'propia': 'deuda', 'yo_debo': 'deuda',
}

FRECUENCIAS = {'mensual': 'mensual', 'mes': 'mensual', 'semanal': 'semanal', 'semana': 'semanal'}

MODOS = {
    'plazo_fijo': 'fixed_term', 'fixed_term': 'fixed_term', 'plazo': 'fixed_term',
    'pago_fijo': 'fixed_payment', 'fixed_payment': 'fixed_payment', 'pago': 'fixed_payment',
}

PLATAFORMAS = {
    'cetesdirecto': 'cetesdirecto', 'cetes': 'cetesdirecto', 'cetes_directo': 'cetesdirecto',
    'briq': 'briq', 'briqmx': 'briq', 'briq_mx': 'briq',
    'otra': 'otra', 'otro': 'otra',
}

TIPOS_INSTRUMENTO = {
    'descuento': 'descuento', 'a_descuento': 'descuento', 'cetes': 'descuento',
    'tasa_fija': 'tasa_fija', 'fija': 'tasa_fija', 'plazo': 'tasa_fija',
    'fondo': 'fondo', 'fondos': 'fondo',
}


def _fila_movimiento_prestamo(fila):
    return {
        'fecha': a_fecha(fila.get('fecha'), 'fecha'),
        'monto': a_decimal(fila.get('monto'), 'monto', minimo=Decimal('0.01')),
        'tipo': a_opcion(fila.get('tipo'), 'tipo', TIPOS_MOVIMIENTO_PRESTAMO, defecto='pago'),
        'descripcion': texto_opcional(fila.get('descripcion')),
    }


def _fila_movimiento_inversion(fila):
    return {
        'fecha': a_fecha(fila.get('fecha'), 'fecha'),
        'monto': a_decimal(fila.get('monto'), 'monto', minimo=Decimal('0.01')),
        'tipo': a_opcion(fila.get('tipo'), 'tipo', TIPOS_MOVIMIENTO_INVERSION),
        'descripcion': texto_opcional(fila.get('descripcion')),
    }


def _fila_prestamo(fila):
    modo = a_opcion(fila.get('modo'), 'modo', MODOS, defecto='fixed_payment')
    datos = {
        'rol': a_opcion(fila.get('rol'), 'rol', ROLES, defecto='prestamo'),
        'concepto': texto_opcional(fila.get('concepto')),
        'nombre_cliente': texto_opcional(fila.get('contraparte') or fila.get('nombre'), 200),
        'telefono': texto_opcional(fila.get('telefono'), 20),
        'monto_original': a_decimal(fila.get('monto'), 'monto', minimo=Decimal('0.01')),
        'tasa_interes_anual': a_decimal(fila.get('tasa'), 'tasa', minimo=Decimal('0')),
        'tipo_pago': a_opcion(fila.get('frecuencia'), 'frecuencia', FRECUENCIAS, defecto='mensual'),
        'fecha_inicio': a_fecha(fila.get('fecha_inicio') or fila.get('fecha'), 'fecha_inicio'),
        'modo': modo,
    }
    if not datos['nombre_cliente']:
        raise ValueError('Falta contraparte (el cliente o el acreedor).')

    plazo = fila.get('plazo')
    pago = fila.get('pago') or fila.get('cuota')
    datos['plazo_meses'] = a_entero(plazo, 'plazo', minimo=1) if str(plazo or '').strip() else None
    datos['pago_mensual'] = (a_decimal(pago, 'pago', minimo=Decimal('0.01'))
                             if str(pago or '').strip() else None)

    if modo == 'fixed_term' and not datos['plazo_meses']:
        raise ValueError('En modo Plazo Fijo hace falta la columna plazo.')
    if modo == 'fixed_payment' and not datos['pago_mensual']:
        raise ValueError('En modo Pago Fijo hace falta la columna pago.')
    return datos


def _fila_inversion(fila):
    tipo = a_opcion(fila.get('tipo'), 'tipo', TIPOS_INSTRUMENTO, defecto='descuento')
    datos = {
        'plataforma': a_opcion(fila.get('plataforma'), 'plataforma', PLATAFORMAS, defecto='otra'),
        'nombre': texto_opcional(fila.get('nombre') or fila.get('instrumento'), 200),
        'tipo': tipo,
        'monto_invertido': a_decimal(fila.get('monto'), 'monto', minimo=Decimal('0.01')),
        'fecha_compra': a_fecha(fila.get('fecha_compra') or fila.get('fecha'), 'fecha_compra'),
        'notas': texto_opcional(fila.get('notas'), 500),
    }
    if not datos['nombre']:
        raise ValueError('Falta nombre del instrumento.')

    valor = fila.get('valor') or fila.get('valor_actual')
    datos['valor_manual'] = (a_decimal(valor, 'valor', minimo=Decimal('0'))
                             if str(valor or '').strip() else None)

    if tipo == 'fondo':
        datos['tasa_anual'] = Decimal('0')
        datos['plazo_dias'] = 0
        datos['base_dias'] = 360
        if datos['valor_manual'] is None:
            raise ValueError('Un fondo no se proyecta: hace falta la columna valor.')
    else:
        datos['tasa_anual'] = a_decimal(fila.get('tasa'), 'tasa', minimo=Decimal('0'))
        datos['plazo_dias'] = a_entero(fila.get('plazo_dias') or fila.get('plazo'),
                                       'plazo_dias', minimo=1)
        base = str(fila.get('base_dias') or '').strip()
        datos['base_dias'] = a_entero(base, 'base_dias') if base else (
            360 if tipo == 'descuento' else 365)
        if datos['base_dias'] not in (360, 365):
            raise ValueError('base_dias sólo puede ser 360 o 365.')
    return datos


IMPORTACIONES = {
    'movimientos_prestamo': {
        'etiqueta': 'Pagos y movimientos de un préstamo',
        'destino': 'prestamo',
        'columnas': 'fecha, monto, tipo (pago|incremento), descripcion',
        'ejemplo': 'fecha,monto,tipo,descripcion\n2025-04-18,3975.00,pago,Semana 1',
        'parser': _fila_movimiento_prestamo,
    },
    'movimientos_inversion': {
        'etiqueta': 'Movimientos de una inversión',
        'destino': 'inversion',
        'columnas': 'fecha, monto, tipo (aportacion|retiro|rendimiento), descripcion',
        'ejemplo': 'fecha,monto,tipo,descripcion\n2025-10-01,10312.42,aportacion,Pago Mensual',
        'parser': _fila_movimiento_inversion,
    },
    'prestamos': {
        'etiqueta': 'Préstamos y deudas',
        'destino': None,
        'columnas': ('rol (prestamo|deuda), contraparte, concepto, telefono, monto, tasa, '
                     'frecuencia (mensual|semanal), fecha_inicio, modo (plazo_fijo|pago_fijo), '
                     'plazo, pago'),
        'ejemplo': ('rol,contraparte,concepto,monto,tasa,frecuencia,fecha_inicio,modo,pago\n'
                    'prestamo,Oscar,Placas,408000,21,semanal,2025-04-03,pago_fijo,3975'),
        'parser': _fila_prestamo,
    },
    'inversiones': {
        'etiqueta': 'Posiciones del portafolio',
        'destino': None,
        'columnas': ('plataforma (cetesdirecto|briq|otra), nombre, tipo (descuento|tasa_fija|fondo), '
                     'monto, fecha_compra, tasa, plazo_dias, base_dias, valor, notas'),
        'ejemplo': ('plataforma,nombre,tipo,monto,fecha_compra,tasa,plazo_dias\n'
                    'cetesdirecto,CETES 28 días,descuento,50000,2026-07-20,9.75,28'),
        'parser': _fila_inversion,
    },
}


def procesar(filas, tipo):
    """
    Valida cada fila y devuelve (validas, errores).

    `validas` son dicts listos para crear el modelo, con `_linea` para poder
    señalar la fila del archivo en la vista previa. `errores` lleva la línea y
    el motivo, para que el usuario corrija sin adivinar.
    """
    parser = IMPORTACIONES[tipo]['parser']
    validas, errores = [], []
    for fila in filas:
        linea = fila.get('_linea', '?')
        try:
            datos = parser(fila)
            # 'linea' sin guion bajo: las plantillas de Django rechazan los
            # nombres que empiezan por '_'.
            datos['linea'] = linea
            validas.append(datos)
        except ValueError as exc:
            errores.append({'linea': linea, 'motivo': str(exc)})
        except Exception as exc:
            errores.append({'linea': linea, 'motivo': f'Fila ilegible: {exc}'})
    return validas, errores
